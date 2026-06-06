# -*- coding: utf-8 -*-
"""
sdl_dwr_step1_generator.py  —  v1.1.0  (bar-fix edition)
Generate Step 1 SDL Daily Work Report DOCX from the daily Excel input file.

Changes from v1.0.0
-------------------
* Removed paragraph-anchored floating "Rectangle 3" bars from document.xml
  (they drifted and disappeared on the last page).
* Injected a single persistent full-page purple sidebar into header1.xml so
  Word repeats it on every page, top-to-bottom, page 1 to last.
* ZIP output now preserves original per-item compress_type to prevent Word corruption.

Required packages : openpyxl, lxml
Expected template  : Daily Work Report Step 1 Template - enTop v1.0.0.docx
Public function    : generate_dwr_step1(excel_path, template_path=None, output_dir=None) -> str
"""
from __future__ import annotations
import html, os, re, shutil, zipfile
from copy import deepcopy
from datetime import date, datetime
from typing import Any, Dict, List
from lxml import etree
from openpyxl import load_workbook

BASE_DIR           = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE   = os.path.join(BASE_DIR, "Daily Work Report Step 1 Template - enTop v1.0.0.docx")
DEFAULT_OUTPUT_DIR = os.path.join(BASE_DIR, "generated_reports")

STATUS_LABELS = {"IP": "In Progress", "NA": "Not attempted", "GU": "Given up"}
W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS   = {"w": W_NS}

# ---------------------------------------------------------------------------
# Purple sidebar — replaces the template's paragraph-anchored Rectangle 3 bars.
#   * Lives in header1.xml  → Word repeats it on every page automatically.
#   * positionH relativeFrom="page", posOffset=7620 EMU  → left edge, every page.
#   * positionV relativeFrom="page", posOffset=0          → top of page.
#   * cy = 10 692 130 EMU  (full A4 height)               → top-to-bottom bar.
#   * cx = 228 600 EMU                                    → same width as original.
#   * behindDoc="1"                                       → behind all content.
#   * accent1 + lumMod 50000                              → same dark purple.
# ---------------------------------------------------------------------------
_SIDEBAR_XML = (
    '<w:r xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
    '<w:rPr><w:noProof/></w:rPr>'
    '<w:pict w14:anchorId="DWRLEFTBAR">'
    '<v:rect id="_x0000_s4097" '
    'style="position:absolute;margin-left:0;margin-top:0;width:18pt;height:842pt;z-index:251660288;visibility:visible;'
    'mso-wrap-style:square;mso-width-percent:0;mso-height-percent:0;'
    'mso-wrap-distance-left:0;mso-wrap-distance-top:0;mso-wrap-distance-right:0;mso-wrap-distance-bottom:0;'
    'mso-position-horizontal:absolute;mso-position-horizontal-relative:page;'
    'mso-position-vertical:absolute;mso-position-vertical-relative:page;'
    'mso-width-percent:0;mso-height-percent:0;v-text-anchor:top" '
    'fillcolor="#6f55d7" stroked="f">'
    '<w10:wrap anchorx="page" anchory="page"/><w10:anchorlock/>'
    '</v:rect></w:pict></w:r>'
)


# ── Type helpers ──────────────────────────────────────────────────────────────

def _safe_float(v: Any, d: float = 0.0) -> float:
    if v is None or v == "": return d
    try: return float(v)
    except (TypeError, ValueError): return d

def _safe_int(v: Any, d: int = 0) -> int:
    if v is None or v == "": return d
    try: return int(round(float(v)))
    except (TypeError, ValueError): return d

def _fmt_int(v):      return str(_safe_int(v))
def _fmt_hours(v):    return f"{_safe_float(v):.2f}"
def _fmt_currency(v): return f"${_safe_float(v):,.2f}"

def _fmt_pct(v):
    n = _safe_float(v)
    return f"{n*100:.2f}%" if abs(n) <= 10 else f"{n:.2f}%"

def _norm_job_no(r):
    if r is None: return ""
    if isinstance(r, (int, float)): return str(int(r))
    return str(r).strip()

def _norm_cust(r):   return " ".join(str(r).split()) if r else ""
def _norm_status(r): return str(r).strip().upper() if r is not None else ""
def _norm_grp(r):    return str(r).strip() if r is not None else ""

def _excel_date(v: Any) -> date:
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try: return datetime.strptime(str(v)[:10], fmt).date()
        except Exception: pass
    return date.today()


# ── Excel reader ──────────────────────────────────────────────────────────────

def _read_excel_data(excel_path: str) -> Dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True)
    if "Summary" not in wb.sheetnames or "Production_Center" not in wb.sheetnames:
        raise ValueError("Workbook must contain 'Summary' and 'Production_Center' sheets.")
    ws_sum, ws_pc = wb["Summary"], wb["Production_Center"]

    dwr_no      = str(ws_sum["B1"].value).strip()
    report_date = _excel_date(ws_sum["C1"].value)

    jobs: List[Dict[str, Any]] = []
    for r in range(8, min(ws_pc.max_row, 108) + 1):
        jtype  = str(ws_pc.cell(r, 4).value or "").strip().upper()
        job_no = _norm_job_no(ws_pc.cell(r, 5).value)
        if not jtype and not job_no: continue
        grp = _norm_grp(ws_pc.cell(r, 15).value)
        if grp.lower() == "continued": continue
        jobs.append({
            "type":     jtype,
            "job_no":   job_no,
            "customer": _norm_cust(ws_pc.cell(r, 12).value),
            "rm_time":  _safe_float(ws_pc.cell(r, 21).value),
            "status":   _norm_status(ws_pc.cell(r, 25).value),
        })

    done = [j for j in jobs if j["status"] == "C"]
    pend = [j for j in jobs if j["status"] in {"IP", "NA", "GU"}]

    c115 = _safe_int(ws_sum["C115"].value)
    d115 = _safe_int(ws_sum["D115"].value)
    e115 = _safe_int(ws_sum["E115"].value)
    g115 = _safe_float(ws_sum["G115"].value)
    h115 = _safe_float(ws_sum["H115"].value)
    i115 = _safe_float(ws_sum["I115"].value)
    inc_hrs = (h115 + i115) or sum(j["rm_time"] for j in pend if j["status"] == "IP")

    return {
        "dwr_no":           dwr_no,
        "date":             report_date,
        "completed":        done,
        "incomplete":       pend,
        "q_summary":        [ws_sum.cell(13, c).value for c in range(1, 10)],
        "o_summary":        [ws_sum["J13"].value] + [ws_sum.cell(13, c).value for c in range(11, 19)],
        "total_summary":    [ws_sum.cell(19, c).value for c in range(1, 11)],
        "resource_rows":    [[ws_sum.cell(r, c).value for c in range(3, 10)] for r in range(3, 7)],
        "completed_count":  c115 or len(done),
        "completed_hours":  g115 or sum(j["rm_time"] for j in done),
        "incomplete_count": (d115 + e115) or len(pend),
        "incomplete_hours": inc_hrs,
    }


# ── XML helpers ───────────────────────────────────────────────────────────────

def _parse_xml(data: bytes) -> etree._Element:
    return etree.fromstring(data, etree.XMLParser(remove_blank_text=False, recover=False))

def _serialize_xml(root: etree._Element) -> bytes:
    return etree.tostring(root, encoding="UTF-8", xml_declaration=True, standalone=False)

def _patch_text(s: str, repls: Dict[str, str]) -> str:
    for k, v in repls.items(): s = s.replace(k, v)
    return s

def _make_run(rpr, text: str, bold: bool) -> etree._Element:
    r = etree.Element(f"{{{W_NS}}}r")
    if rpr is not None:
        nr = deepcopy(rpr)
        for tag in (f"{{{W_NS}}}b", f"{{{W_NS}}}bCs"):
            el = nr.find(tag)
            if el is not None: nr.remove(el)
        if bold:
            nr.insert(0, etree.Element(f"{{{W_NS}}}bCs"))
            nr.insert(0, etree.Element(f"{{{W_NS}}}b"))
        r.append(nr)
    t = etree.SubElement(r, f"{{{W_NS}}}t")
    t.text = text
    if text.startswith(" ") or text.endswith(" "):
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    return r

def _footer_label(tc, kw: str, n: int):
    p = tc.find(f"{{{W_NS}}}p") or etree.SubElement(tc, f"{{{W_NS}}}p")
    runs = p.xpath("./w:r", namespaces=NS)
    rp = rb = None
    for run in runs:
        rpr = run.find(f"{{{W_NS}}}rPr")
        if rpr is None: continue
        if rpr.find(f"{{{W_NS}}}b") is not None:
            if rb is None: rb = rpr
        else:
            if rp is None: rp = rpr
    rp = rp or rb; rb = rb or rp
    for run in runs: p.remove(run)
    for text, bold in [("Total ",False),(kw,True),(" PC jobs ",False),(f"{n} Nos.",True)]:
        p.append(_make_run(rb if bold else rp, text, bold))

def _set_tc(tc, text: Any):
    val = "" if text is None else str(text)
    ts  = tc.xpath(".//w:t", namespaces=NS)
    if ts:
        ts[0].text = val
        if val.startswith(" ") or val.endswith(" "):
            ts[0].set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        for t in ts[1:]: t.text = ""
        return
    p = tc.find(f"{{{W_NS}}}p") or etree.SubElement(tc, f"{{{W_NS}}}p")
    r = p.find(f"{{{W_NS}}}r")  or etree.SubElement(p, f"{{{W_NS}}}r")
    t = etree.SubElement(r, f"{{{W_NS}}}t")
    t.text = val
    if val.startswith(" ") or val.endswith(" "):
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")

def _set_row(row, vals):
    cells = row.xpath("./w:tc", namespaces=NS)
    for i, v in enumerate(vals):
        if i < len(cells): _set_tc(cells[i], v)

def _cell_text(tc) -> str:
    return "".join(t.text or "" for t in tc.xpath(".//w:t", namespaces=NS)).strip()

def _set_tc_fill(tc, fill_hex: str) -> None:
    """Apply a solid Word table-cell fill colour without changing borders/text."""
    fill = fill_hex.strip().lstrip("#").upper()
    tcPr = tc.find(f"{{{W_NS}}}tcPr")
    if tcPr is None:
        tcPr = etree.Element(f"{{{W_NS}}}tcPr")
        tc.insert(0, tcPr)
    for shd in list(tcPr.findall(f"{{{W_NS}}}shd")):
        tcPr.remove(shd)
    shd = etree.Element(f"{{{W_NS}}}shd")
    shd.set(f"{{{W_NS}}}val", "clear")
    shd.set(f"{{{W_NS}}}color", "auto")
    shd.set(f"{{{W_NS}}}fill", fill)
    tcPr.append(shd)

def _set_tc_text_color(tc, color_hex: str) -> None:
    """Apply a font colour to all text runs in a table cell."""
    color = color_hex.strip().lstrip("#").upper()
    runs = tc.xpath(".//w:r", namespaces=NS)
    if not runs:
        return
    for run in runs:
        rPr = run.find(f"{{{W_NS}}}rPr")
        if rPr is None:
            rPr = etree.Element(f"{{{W_NS}}}rPr")
            run.insert(0, rPr)
        for old in list(rPr.findall(f"{{{W_NS}}}color")):
            rPr.remove(old)
        color_el = etree.Element(f"{{{W_NS}}}color")
        color_el.set(f"{{{W_NS}}}val", color)
        rPr.append(color_el)

def _apply_status_coloring(row) -> None:
    """Colour only the 'Not attempted' status text in incomplete-task rows."""
    cells = row.xpath("./w:tc", namespaces=NS)
    if len(cells) < 5:
        return
    status_value = _cell_text(cells[-1]).strip().lower()
    if status_value == "not attempted":
        _set_tc_text_color(cells[-1], "F26E61")

def _apply_type_shading(row) -> None:
    """Shade Type cells: O = #EAF7FC and Q = #FFFFE7."""
    cells = row.xpath("./w:tc", namespaces=NS)
    if not cells:
        return
    type_value = _cell_text(cells[0]).upper()
    if type_value == "O":
        _set_tc_fill(cells[0], "EAF7FC")
    elif type_value == "Q":
        _set_tc_fill(cells[0], "FFFFE7")

def _rebuild_table(tbl, data_rows, footer_vals, fkw=None, fn=None):
    rows = tbl.xpath("./w:tr", namespaces=NS)
    if len(rows) < 3: return
    dtmpl = deepcopy(rows[1]); ftmpl = deepcopy(rows[-1])
    for r in rows[1:]: tbl.remove(r)
    for vals in data_rows:
        nr = deepcopy(dtmpl); _set_row(nr, vals); _apply_type_shading(nr); _apply_status_coloring(nr); tbl.append(nr)
    fr = deepcopy(ftmpl)
    if fkw and fn is not None:
        cells = fr.xpath("./w:tc", namespaces=NS)
        for i, v in enumerate(footer_vals):
            if i != 2 and i < len(cells): _set_tc(cells[i], v)
        if len(cells) > 2: _footer_label(cells[2], fkw, fn)
    else:
        _set_row(fr, footer_vals)
    tbl.append(fr)

def _drop_blank_para_before(tbl):
    parent = tbl.getparent()
    if parent is None: return
    siblings = list(parent); idx = siblings.index(tbl)
    if idx == 0: return
    prev = siblings[idx - 1]
    try:
        if etree.QName(prev.tag).localname == "p":
            if not any((t.text or "").strip() for t in prev.xpath(".//w:t", namespaces=NS)):
                parent.remove(prev)
    except Exception: pass


# ── Bar-fix functions ─────────────────────────────────────────────────────────

def _is_old_sidebar_run(run) -> bool:
    """Return True for the old floating purple sidebar run only."""
    block = etree.tostring(run, encoding="unicode").lower()
    return (
        ('name="rectangle 3"' in block and ('<wp:docpr' in block or 'docpr' in block))
        or (
            '#6f55d7' in block
            and 'width:18pt' in block
            and ('height:842pt' in block or 'height:843.4pt' in block or 'height:842.0pt' in block)
        )
        or ('w14:anchorid="dwrleftbar"' in block.lower())
    )

def _remove_old_sidebar_runs(root) -> None:
    """Remove old body/header floating sidebar objects without touching logos/tables."""
    doomed_paras = []
    for run in list(root.xpath("//w:r", namespaces=NS)):
        if _is_old_sidebar_run(run):
            parent = run.getparent()
            parent.remove(run)
            if parent is not None and etree.QName(parent.tag).localname == "p":
                txt = ''.join(t.text or '' for t in parent.xpath('.//w:t', namespaces=NS)).strip()
                has_drawing = bool(parent.xpath('.//*[local-name()="drawing" or local-name()="pict"]'))
                has_tbl = bool(parent.xpath('.//*[local-name()="tbl"]'))
                if not txt and not has_drawing and not has_tbl:
                    doomed_paras.append(parent)
    for para in doomed_paras:
        parent = para.getparent()
        if parent is not None:
            parent.remove(para)

def _strip_rectangle3(doc_xml: str) -> str:
    """Remove all old floating sidebar objects safely from document.xml."""
    try:
        root = _parse_xml(doc_xml.encode("utf-8"))
        _remove_old_sidebar_runs(root)
        return _serialize_xml(root).decode("utf-8")
    except Exception:
        # Fallback for damaged XML: remove only self-contained drawing/pict runs.
        doc_xml = re.sub(
            r'<w:r\b[^>]*>.*?name="Rectangle 3".*?</w:r>',
            '', doc_xml, flags=re.DOTALL,
        )
        doc_xml = re.sub(
            r'<w:pict\b[^>]*>.*?fillcolor="#6f55d7".*?</w:pict>',
            '', doc_xml, flags=re.DOTALL | re.IGNORECASE,
        )
        return doc_xml

def _clean_header_sidebar(header_xml: str) -> str:
    """Remove any pre-existing sidebar bar from a header before injecting one stable bar."""
    try:
        root = _parse_xml(header_xml.encode("utf-8"))
        _remove_old_sidebar_runs(root)
        return _serialize_xml(root).decode("utf-8")
    except Exception:
        header_xml = re.sub(r'<w:r\b[^>]*>.*?w14:anchorId="DWRLEFTBAR".*?</w:r>', '', header_xml, flags=re.DOTALL)
        return header_xml

def _inject_sidebar(header_xml: str) -> str:
    """Insert persistent full-page sidebar as its own top-level header paragraph."""
    header_xml = _clean_header_sidebar(header_xml)
    bar_para = (
        '<w:p w14:paraId="8B4F5A21" w14:textId="77777777" '
        'w:rsidR="00A5655E" w:rsidRDefault="00A5655E">'
        '<w:pPr><w:pStyle w:val="Header"/></w:pPr>'
        + _SIDEBAR_XML +
        '</w:p>'
    )
    if 'w14:anchorId="DWRLEFTBAR"' in header_xml:
        return header_xml
    return re.sub(r'(<w:hdr\b[^>]*>)', r'\1' + bar_para, header_xml, count=1)

def _next_relationship_id(rels_xml: str) -> str:
    ids = [int(x) for x in re.findall(r'Id="rId(\d+)"', rels_xml)]
    return f"rId{max(ids + [0]) + 1}"


# ── Core patcher ──────────────────────────────────────────────────────────────

def _patch_document_xml(doc_bytes: bytes, data: Dict[str, Any]) -> bytes:
    raw  = _strip_rectangle3(doc_bytes.decode("utf-8"))
    root = _parse_xml(raw.encode("utf-8"))
    tbls = root.xpath("//w:tbl", namespaces=NS)
    if len(tbls) < 6: raise ValueError("Template must have at least 6 tables.")

    _set_row(tbls[0].xpath("./w:tr", namespaces=NS)[3],
             [f(v) for v,f in zip(data["q_summary"],
              [_fmt_int]*4+[_fmt_hours]*2+[_fmt_int]*2+[_fmt_currency])])
    _set_row(tbls[1].xpath("./w:tr", namespaces=NS)[3],
             [f(v) for v,f in zip(data["o_summary"],
              [_fmt_int]*4+[_fmt_hours]*2+[_fmt_int]*2+[_fmt_currency])])
    _set_row(tbls[2].xpath("./w:tr", namespaces=NS)[3],
             [f(v) for v,f in zip(data["total_summary"],
              [_fmt_int]*6+[_fmt_hours]*3+[_fmt_currency])])

    _rebuild_table(tbls[3],
        [[j["type"],j["job_no"],j["customer"],_fmt_hours(j["rm_time"])] for j in data["completed"]],
        ["","","",_fmt_hours(data["completed_hours"])],
        fkw="completed", fn=data["completed_count"])

    _rebuild_table(tbls[4],
        [[j["type"],j["job_no"],j["customer"],
          _fmt_hours(j["rm_time"]) if j["status"]=="IP" else "-",
          STATUS_LABELS.get(j["status"],j["status"])] for j in data["incomplete"]],
        ["","","",_fmt_hours(data["incomplete_hours"]),"Hours"],
        fkw="incomplete", fn=data["incomplete_count"])
    _drop_blank_para_before(tbls[4])

    res_rows = tbls[5].xpath("./w:tr", namespaces=NS)
    for idx, rv in enumerate(data["resource_rows"], 1):
        if idx < len(res_rows):
            _set_row(res_rows[idx],
                     [f(v) for v,f in zip(rv,
                      [str,_fmt_int,_fmt_int,_fmt_hours,_fmt_hours,_fmt_hours,_fmt_pct])])

    d   = data["date"]; dwr = data["dwr_no"]
    dl  = f"{d.day} {d.strftime('%B %Y')}"
    xml = _patch_text(_serialize_xml(root).decode("utf-8"), {
        "Publish Date": dl, "DWR# 2026-XX": f"DWR# {dwr}",
        "DWR# 2026-90": f"DWR# {dwr}", "21/05/26": d.strftime("%d/%m/%y"),
        "21 May 2026":  dl,
    })
    return xml.encode("utf-8")


def _patch_all_xml_parts(docx_path: str, data: Dict[str, Any]) -> None:
    d     = data["date"]; dwr = data["dwr_no"]
    dl    = f"{d.day} {d.strftime('%B %Y')}"; iso = d.strftime("%Y-%m-%dT00:00:00")
    repls = {
        "Publish Date": dl, "DWR# 2026-XX": f"DWR# {dwr}",
        "DWR# 2026-90": f"DWR# {dwr}", "21/05/26": d.strftime("%d/%m/%y"),
        "21 May 2026":  dl,
    }
    tmp = docx_path + ".tmp"

    with zipfile.ZipFile(docx_path, "r") as zin:
        names = set(zin.namelist())
        original_doc_xml = zin.read("word/document.xml").decode("utf-8")
        rels_xml_original = zin.read("word/_rels/document.xml.rels").decode("utf-8")
        content_types_original = zin.read("[Content_Types].xml").decode("utf-8")
        header1_original = zin.read("word/header1.xml").decode("utf-8") if "word/header1.xml" in names else None

        needs_first_header = (
            '<w:titlePg' in original_doc_xml
            and '<w:headerReference w:type="first"' not in original_doc_xml
            and header1_original is not None
        )
        new_first_header_rid = _next_relationship_id(rels_xml_original) if needs_first_header else None
        header2_content = None
        if needs_first_header:
            header2_content = _inject_sidebar(_patch_text(header1_original, repls)).encode("utf-8")

        with zipfile.ZipFile(tmp, "w") as zout:
            wrote_header2 = False
            for item in zin.infolist():
                content = zin.read(item.filename)

                if item.filename == "word/document.xml":
                    content = _patch_document_xml(content, data)
                    if needs_first_header and new_first_header_rid:
                        t = content.decode("utf-8")
                        t = re.sub(
                            r'(<w:headerReference\s+w:type="default"\s+r:id="[^"]+"\s*/>)',
                            r'\1' + f'<w:headerReference w:type="first" r:id="{new_first_header_rid}"/>',
                            t,
                            count=1,
                        )
                        content = t.encode("utf-8")

                elif item.filename == "word/header1.xml":
                    try:
                        content = _inject_sidebar(
                            _patch_text(content.decode("utf-8"), repls)
                        ).encode("utf-8")
                    except UnicodeDecodeError:
                        pass

                elif item.filename == "word/header2.xml":
                    try:
                        content = _inject_sidebar(
                            _patch_text(content.decode("utf-8"), repls)
                        ).encode("utf-8")
                        wrote_header2 = True
                    except UnicodeDecodeError:
                        pass

                elif item.filename == "word/_rels/document.xml.rels" and needs_first_header and new_first_header_rid:
                    t = content.decode("utf-8")
                    if f'Id="{new_first_header_rid}"' not in t:
                        t = t.replace(
                            "</Relationships>",
                            f'<Relationship Id="{new_first_header_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/header" Target="header2.xml"/></Relationships>'
                        )
                    content = t.encode("utf-8")

                elif item.filename == "[Content_Types].xml" and needs_first_header:
                    t = content.decode("utf-8")
                    override = '<Override PartName="/word/header2.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.header+xml"/>'
                    if 'PartName="/word/header2.xml"' not in t:
                        t = t.replace("</Types>", override + "</Types>")
                    content = t.encode("utf-8")

                elif item.filename.endswith(".xml") and (
                    item.filename.startswith("word/") or
                    item.filename.startswith("docProps/") or
                    item.filename.startswith("customXml/")):
                    try:
                        t = _patch_text(content.decode("utf-8"), repls)
                        if item.filename == "docProps/core.xml":
                            t = re.sub(r"<cp:contentStatus>.*?</cp:contentStatus>",
                                       f"<cp:contentStatus>{html.escape(dwr)}</cp:contentStatus>", t)
                        if item.filename.startswith("customXml/"):
                            t = re.sub(r"<PublishDate>.*?</PublishDate>",
                                       f"<PublishDate>{iso}</PublishDate>", t)
                        t = re.sub(r'<w:date w:fullDate="[^"]*"',
                                   f'<w:date w:fullDate="{iso}Z"', t)
                        content = t.encode("utf-8")
                    except UnicodeDecodeError:
                        pass

                # Preserve original compress_type — critical for Word compatibility
                zout.writestr(item, content, compress_type=item.compress_type)

            if needs_first_header and header2_content is not None and not wrote_header2:
                zout.writestr("word/header2.xml", header2_content, compress_type=zipfile.ZIP_DEFLATED)

    os.replace(tmp, docx_path)


# ── Public API ────────────────────────────────────────────────────────────────

def generate_dwr_step1(
    excel_path:    str,
    template_path: str | None = None,
    output_dir:    str | None = None,
) -> str:
    """Generate Step 1 DWR DOCX and return the output file path."""
    template_path = template_path or DEFAULT_TEMPLATE
    output_dir    = output_dir    or DEFAULT_OUTPUT_DIR
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")
    os.makedirs(output_dir, exist_ok=True)
    data = _read_excel_data(excel_path)
    d    = data["date"]
    out  = os.path.join(output_dir,
                        f"DWR_{data['dwr_no']}_{d.strftime('%d-%b-%Y')}-Step_1.docx")
    shutil.copyfile(template_path, out)
    _patch_all_xml_parts(out, data)
    return out


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Generate SDL DWR Step 1 DOCX.")
    ap.add_argument("excel_path")
    ap.add_argument("--template",   default=None)
    ap.add_argument("--output-dir", default=None)
    args = ap.parse_args()
    print(generate_dwr_step1(args.excel_path,
                              template_path=args.template,
                              output_dir=args.output_dir))