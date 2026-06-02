# -*- coding: utf-8 -*-
"""
sdl_dwr_step4_generator.py

Generate Step 4 SDL Daily Work Report DOCX from the daily Excel file.
Designed for a small Flask app on Namecheap/cPanel shared hosting.

This version patches the DOCX package XML directly instead of opening/saving
with python-docx. That preserves the Step 4 template's header/footer drawing
objects, including the purple vertical page bar.

Required package:
    openpyxl

Expected server files in the same folder as this script/app:
    Daily Work Report Step 4 Template - enTop v1.0.0.docx

Public function:
    generate_dwr_step4(excel_path, template_path=None, output_dir=None) -> str
"""

from __future__ import annotations

import os
import re
import zipfile
from copy import deepcopy
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple
from xml.sax.saxutils import escape

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(BASE_DIR, "Daily Work Report Step 4 Template - enTop v1.0.0.docx")
DEFAULT_OUTPUT_DIR = os.path.join(BASE_DIR, "generated_reports")

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"
try:
    from lxml import etree
except Exception as exc:  # pragma: no cover - lxml normally exists with python-docx installs
    etree = None


def _safe_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def _normalize_job_no(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, (int, float)):
        return str(int(raw))
    return str(raw).strip()


def _normalize_customer(raw: Any) -> str:
    return " ".join(str(raw).split()) if raw else ""


def _normalize_status(raw: Any) -> str:
    return str(raw).strip().upper() if raw is not None else ""


def _normalize_grp(raw: Any) -> str:
    return str(raw).strip() if raw is not None else ""


def _excel_serial_or_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except Exception:
        return date.today()


def _fmt_currency(value: float, decimals: int = 0) -> str:
    if decimals <= 0:
        return f"${value:,.0f}"
    return f"${value:,.{decimals}f}"


def _fmt_hours(value: float) -> str:
    return f"{value:.2f}"


def _fmt_pct(value: float, decimals: int = 1) -> str:
    return f"{value:.{decimals}f}%"


def _rate(job: Dict[str, Any]) -> float:
    t = _safe_float(job.get("rm_time"))
    return _safe_float(job.get("value")) / t if t > 0 else 0.0


def _efficiency_text(job: Dict[str, Any]) -> str:
    value = _safe_float(job.get("value"))
    time = _safe_float(job.get("rm_time"))
    return f"{job.get('type', '')} {job.get('job_no', '')} ({_fmt_currency(value)}/{_fmt_hours(time)}h = {_fmt_currency(_rate(job))}/hr)"


def _plural(n: int, singular: str, plural: Optional[str] = None) -> str:
    if n == 1:
        return f"{n} {singular}"
    return f"{n} {plural or singular + 's'}"


def _read_excel_data(excel_path: str) -> Dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True)
    if "Summary" not in wb.sheetnames or "Production_Center" not in wb.sheetnames:
        raise ValueError("Input workbook must contain 'Summary' and 'Production_Center' sheets.")

    ws_sum = wb["Summary"]
    ws_pc = wb["Production_Center"]

    dwr_no = str(ws_sum["B1"].value).strip()
    report_date = _excel_serial_or_date(ws_sum["C1"].value)

    jobs: List[Dict[str, Any]] = []
    for r in range(8, min(ws_pc.max_row, 108) + 1):
        job_type = str(ws_pc.cell(r, 4).value or "").strip().upper()
        job_no = _normalize_job_no(ws_pc.cell(r, 5).value)
        if not job_type and not job_no:
            continue

        grp_work = _normalize_grp(ws_pc.cell(r, 15).value)
        if grp_work.lower() == "continued":
            continue

        jobs.append({
            "row": r,
            "type": job_type,
            "job_no": job_no,
            "value": _safe_float(ws_pc.cell(r, 11).value),
            "customer": _normalize_customer(ws_pc.cell(r, 12).value),
            "grp_work": grp_work,
            "rm_time": _safe_float(ws_pc.cell(r, 21).value),
            "status": _normalize_status(ws_pc.cell(r, 25).value),
        })

    active_all = [j for j in jobs if j["status"] in {"C", "IP"}]
    active_jobs = [j for j in active_all if j["rm_time"] > 0 and j["value"] > 0]
    active_o = [j for j in active_jobs if j["type"] == "O"]
    active_q = [j for j in active_jobs if j["type"] == "Q"]
    active_all_o = [j for j in active_all if j["type"] == "O"]
    active_all_q = [j for j in active_all if j["type"] == "Q"]
    incomplete = [j for j in jobs if j["status"] in {"IP", "NA", "GU"}]

    actual_hours = _safe_float(ws_sum["H19"].value)
    if actual_hours <= 0:
        actual_hours = sum(j["rm_time"] for j in active_all)

    pending_count = _safe_int(ws_sum["E115"].value)
    ip_count = _safe_int(ws_sum["D115"].value)
    if pending_count <= 0:
        pending_count = len([j for j in incomplete if j["status"] in {"NA", "GU"}])
    if ip_count <= 0:
        ip_count = len([j for j in incomplete if j["status"] == "IP"])

    return {
        "dwr_no": dwr_no,
        "date": report_date,
        "jobs": jobs,
        "active_all": active_all,
        "active_jobs": active_jobs,
        "active_o": active_o,
        "active_q": active_q,
        "active_all_o": active_all_o,
        "active_all_q": active_all_q,
        "incomplete": incomplete,
        "actual_hours": actual_hours,
        "pending_count": pending_count,
        "ip_count": ip_count,
        "ord_val": sum(j["value"] for j in active_o),
        "q_val": sum(j["value"] for j in active_q),
        "total_val": sum(j["value"] for j in active_jobs),
        "ord_hrs": sum(j["rm_time"] for j in active_o),
        "q_hrs": sum(j["rm_time"] for j in active_q),
    }


def _top_customer_by_value(jobs: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for j in jobs:
        customer = j["customer"] or "UNKNOWN CUSTOMER"
        if customer not in grouped:
            grouped[customer] = {"customer": customer, "value": 0.0, "count": 0, "time": 0.0, "jobs": []}
        grouped[customer]["value"] += j["value"]
        grouped[customer]["count"] += 1
        grouped[customer]["time"] += j["rm_time"]
        grouped[customer]["jobs"].append(j)
    if not grouped:
        return None
    return sorted(grouped.values(), key=lambda x: (-x["value"], x["customer"]))[0]


def _time_heavy_order(data: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], str]:
    active_o = data["active_o"]
    if not active_o:
        return None, "No active order jobs were available for time-heavy order review."

    ip_orders = [j for j in active_o if j["status"] == "IP"]
    if ip_orders:
        selected = sorted(ip_orders, key=lambda j: (-j["rm_time"], _rate(j), j["job_no"]))[0]
    else:
        selected = sorted(active_o, key=lambda j: (-j["rm_time"], _rate(j), j["job_no"]))[0]

    same_customer = [j for j in active_o if j["customer"] == selected["customer"]]
    total_time = sum(j["rm_time"] for j in same_customer)
    total_value = sum(j["value"] for j in same_customer)
    avg_rate = total_value / total_time if total_time > 0 else 0.0
    status_note = " still in progress" if selected["status"] == "IP" else " was the largest order time consumer"
    detail = (
        f"{selected['customer']} = Time-Heavy Orders: {_plural(len(same_customer), 'order')} consuming "
        f"{_fmt_hours(total_time)} RM hours for {_fmt_currency(total_value)} combined. "
        f"{selected['type']} {selected['job_no']} ({selected['status'] or 'C'}, {_fmt_hours(selected['rm_time'])} hrs){status_note}. "
        f"Combined yield is approximately {_fmt_currency(avg_rate)}/hr, so monitor multi-part or high-effort order processing."
    )
    return selected, detail


def _build_observations(data: Dict[str, Any]) -> List[str]:
    active_jobs = data["active_jobs"]
    active_o = data["active_o"]
    active_q = data["active_q"]
    total_val = data["total_val"]

    observations: List[str] = []

    leader = _top_customer_by_value(active_jobs)
    if leader and total_val > 0:
        share = leader["value"] / total_val * 100
        best_job = max(leader["jobs"], key=_rate)
        label = "Revenue Anchor" if share >= 85 else "Revenue Leader"
        job_type_word = "orders" if all(j["type"] == "O" for j in leader["jobs"]) else "jobs"
        observations.append(
            f"{leader['customer']} = {label} ({_fmt_pct(share)} of total active value): "
            f"{_plural(leader['count'], job_type_word[:-1] if job_type_word.endswith('s') else job_type_word)} totalling "
            f"{_fmt_currency(leader['value'])}. Best job: {_efficiency_text(best_job)}. "
            f"{leader['customer']} accounts for {_fmt_pct(share)} of all active value today."
        )
    else:
        observations.append("No active revenue leader could be identified from the available active jobs.")

    best_quote = max(active_q, key=_rate) if active_q else None
    if best_quote:
        observations.append(
            f"{best_quote['customer']} = Quote of the Day ({_fmt_currency(best_quote['value'])} in {_fmt_hours(best_quote['rm_time'])} hrs): "
            f"Q {best_quote['job_no']} is the highest-efficiency active quote. At {_fmt_currency(_rate(best_quote))}/hr, "
            f"it is a strong efficiency result. Watch for conversion to order."
        )
    else:
        observations.append("Quote of the Day: no active quote with positive value and RM time was available for efficiency ranking.")

    _, time_heavy_text = _time_heavy_order(data)
    observations.append(time_heavy_text)

    q_customers = {j["customer"] for j in data["active_all_q"] if j["customer"]}
    o_customers = {j["customer"] for j in data["active_all_o"] if j["customer"]}
    both_customers = q_customers.intersection(o_customers)
    observations.append(
        f"Customer Mix: {_plural(len(q_customers), 'unique customer')} were quoted (Q), "
        f"{_plural(len(o_customers), 'unique customer')} placed orders (O), and "
        f"{_plural(len(both_customers), 'customer')} appeared in both quotation and order activity."
    )

    actual_hours = data["actual_hours"]
    load_pct = actual_hours / 40 * 100 if actual_hours else 0.0
    pending_count = data["pending_count"]
    ip_count = data["ip_count"]
    observations.append(
        f"Actual {_fmt_hours(actual_hours)} hrs vs 40-hr baseline = {_fmt_pct(load_pct)} load. "
        f"{pending_count} jobs left Pending/Not Attempted (NA/GU), {ip_count} still In Progress (IP). "
        f"Any backlog on high-value customers represents significant deferred revenue risk."
    )

    best_order = max(active_o, key=_rate) if active_o else None
    if best_order:
        assert best_order["type"] == "O"
        assert round(best_order["value"] / best_order["rm_time"], 2) == round(_rate(best_order), 2)
    if best_quote:
        assert best_quote["type"] == "Q"
        assert round(best_quote["value"] / best_quote["rm_time"], 2) == round(_rate(best_quote), 2)

    return observations


def _output_name(dwr_no: str, report_date: date) -> str:
    return f"DWR_{dwr_no}_{report_date.strftime('%d-%b-%Y')}-Step_4.docx"


def _w(tag: str) -> str:
    return f"{{{W_NS}}}{tag}"


def _text_of_paragraph(p) -> str:
    return "".join(t.text or "" for t in p.xpath('.//w:t', namespaces={'w': W_NS}))


def _first_nonempty_rpr(p):
    for r in p.xpath('./w:r', namespaces={'w': W_NS}):
        if "".join(t.text or "" for t in r.xpath('.//w:t', namespaces={'w': W_NS})).strip():
            rpr = r.find(_w('rPr'))
            if rpr is not None:
                return deepcopy(rpr)
    r = p.find(_w('r'))
    if r is not None:
        rpr = r.find(_w('rPr'))
        if rpr is not None:
            return deepcopy(rpr)
    return None


def _make_run(text: str, base_rpr=None, bold: Optional[bool] = None):
    r = etree.Element(_w('r'))
    if base_rpr is not None:
        rpr = deepcopy(base_rpr)
    else:
        rpr = etree.Element(_w('rPr'))
    if bold is not None:
        # Remove existing bold tags, then add only when requested.
        for b in list(rpr.findall(_w('b'))):
            rpr.remove(b)
        if bold:
            etree.SubElement(rpr, _w('b'))
    if len(rpr):
        r.append(rpr)
    t = etree.SubElement(r, _w('t'))
    if text.startswith(' ') or text.endswith(' '):
        t.set(XML_SPACE, 'preserve')
    t.text = text
    return r


def _replace_para_text_keep_style(p, text: str, bold_prefix: bool = True) -> None:
    base_rpr = _first_nonempty_rpr(p)
    # Remove runs/hyperlinks only; keep paragraph properties, numbering, tabs, bookmarks, etc.
    for child in list(p):
        if child.tag in {_w('r'), _w('hyperlink')}:
            p.remove(child)

    if bold_prefix and ':' in text:
        prefix, suffix = text.split(':', 1)
        p.append(_make_run(prefix + ':', base_rpr, bold=True))
        p.append(_make_run(suffix, base_rpr, bold=False))
    else:
        p.append(_make_run(text, base_rpr, bold=None))


def _replace_observations_in_document_xml(xml_bytes: bytes, observations: List[str]) -> bytes:
    if etree is None:
        raise RuntimeError("lxml is required for Step 4 XML-safe DOCX patching.")
    parser = etree.XMLParser(remove_blank_text=False, recover=True)
    root = etree.fromstring(xml_bytes, parser)
    ns = {'w': W_NS}
    body = root.find('.//w:body', namespaces=ns)
    if body is None:
        raise ValueError("Invalid DOCX: word/document.xml has no body.")
    paragraphs = body.findall(_w('p'))

    heading_idx = None
    for i, p in enumerate(paragraphs):
        if 'Key Observations' in _text_of_paragraph(p):
            heading_idx = i
            break
    if heading_idx is None:
        raise ValueError("Could not find 'Key Observations' heading in Step 4 template.")

    bullet_indexes: List[int] = []
    after_heading = paragraphs[heading_idx + 1:]
    for offset, p in enumerate(after_heading, start=heading_idx + 1):
        txt = _text_of_paragraph(p).strip()
        pstyle = p.xpath('./w:pPr/w:pStyle/@w:val', namespaces=ns)
        # Stop once the next heading/major section begins, but do not stop on blank list placeholders.
        if txt and pstyle and pstyle[0].startswith('Heading'):
            break
        if txt or (pstyle and pstyle[0] in {'ListParagraph', 'ListBullet'}):
            bullet_indexes.append(offset)
        # Keep scanning through list placeholders only; if a non-list blank paragraph appears, it belongs to layout.
        if not txt and not (pstyle and pstyle[0] in {'ListParagraph', 'ListBullet'}):
            break

    populated = [idx for idx in bullet_indexes if _text_of_paragraph(paragraphs[idx]).strip()]
    if not populated:
        raise ValueError("Could not find populated bullet paragraphs under Key Observations.")

    # Ensure enough list paragraphs exist. Use blank list placeholder if present; otherwise clone last bullet.
    while len(bullet_indexes) < len(observations):
        clone = deepcopy(paragraphs[populated[-1]])
        ref = paragraphs[bullet_indexes[-1]] if bullet_indexes else paragraphs[populated[-1]]
        ref.addnext(clone)
        paragraphs = body.findall(_w('p'))
        # Recompute index of inserted clone.
        new_idx = paragraphs.index(clone)
        bullet_indexes.append(new_idx)

    # Replace the first N list paragraphs. Do not delete any other paragraph, because drawings or
    # page-design elements can be anchored to apparently blank XML paragraphs in the template.
    for idx, obs in zip(bullet_indexes[:len(observations)], observations):
        _replace_para_text_keep_style(paragraphs[idx], obs, bold_prefix=True)

    # Clear surplus populated bullet text if there are more template bullets than observations.
    for idx in bullet_indexes[len(observations):]:
        if _text_of_paragraph(paragraphs[idx]).strip():
            _replace_para_text_keep_style(paragraphs[idx], "", bold_prefix=False)

    return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone='yes')


def _replace_xml_text_values(xml_bytes: bytes, replacements: Dict[str, str]) -> bytes:
    # Direct text replacement preserves all drawing/shape XML, including header VML rectangles.
    text = xml_bytes.decode('utf-8', errors='ignore')
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.encode('utf-8')


def _patch_docx_from_template(template_path: str, output_path: str, replacements: Dict[str, str], observations: List[str]) -> None:
    with zipfile.ZipFile(template_path, 'r') as zin, zipfile.ZipFile(output_path, 'w') as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'word/document.xml':
                data = _replace_xml_text_values(data, replacements)
                data = _replace_observations_in_document_xml(data, observations)
            elif item.filename.endswith('.xml'):
                data = _replace_xml_text_values(data, replacements)
            compress_type = zipfile.ZIP_STORED if ('embeddings' in item.filename and item.filename.endswith('.xlsx')) else item.compress_type
            zout.writestr(item, data, compress_type=compress_type)


def generate_dwr_step4(excel_path: str, template_path: str | None = None, output_dir: str | None = None) -> str:
    """Generate Step 4 DWR DOCX and return the output path."""
    template_path = template_path or DEFAULT_TEMPLATE
    output_dir = output_dir or DEFAULT_OUTPUT_DIR
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel input file not found: {excel_path}")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Step 4 template not found: {template_path}")

    os.makedirs(output_dir, exist_ok=True)
    data = _read_excel_data(excel_path)
    observations = _build_observations(data)
    report_date = data["date"]
    dwr_no = data["dwr_no"]

    replacements = {
        "DWR# 2026-90": f"DWR# {dwr_no}",
        "2026-90": dwr_no,
        "21/05/26": report_date.strftime("%d/%m/%y"),
        "21 May 2026": report_date.strftime("%d %B %Y"),
        "Thursday, 21 May 2026": report_date.strftime("%A, %d %B %Y"),
        "[PUBLISH DATE]": report_date.strftime("%d %B %Y"),
    }

    output_path = os.path.join(output_dir, _output_name(dwr_no, report_date))
    _patch_docx_from_template(template_path, output_path, replacements, observations)
    return output_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate SDL DWR Step 4 DOCX.")
    parser.add_argument("excel_path", help="Path to daily Excel input file")
    parser.add_argument("--template", default=None, help="Optional Step 4 template path")
    parser.add_argument("--output-dir", default=None, help="Optional output directory")
    args = parser.parse_args()

    print(generate_dwr_step4(args.excel_path, template_path=args.template, output_dir=args.output_dir))
